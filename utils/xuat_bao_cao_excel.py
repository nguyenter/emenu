from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side


def _nhan_thanh_toan(phuong_thuc):
    if phuong_thuc == 'TIEN_MAT':
        return 'Tiền mặt'
    if phuong_thuc in ('CHUYEN_KHOAN', 'QR'):
        return 'Chuyển khoản / QR'
    return phuong_thuc or '—'


def tao_file_bao_cao_doanh_thu(
    nhan_ky,
    ds_hoa_don,
    tong_doanh_thu,
    tong_hoa_don,
    tien_mat,
    chuyen_khoan_qr,
):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Doanh thu'

    bold = Font(bold=True)
    header_font = Font(bold=True)
    thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    ws['A1'] = 'BÁO CÁO DOANH THU'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    ws['A2'] = f'Kỳ báo cáo: {nhan_ky}'
    ws.merge_cells('A2:G2')

    ws['A4'] = 'Tổng doanh thu'
    ws['B4'] = float(tong_doanh_thu or 0)
    ws['B4'].number_format = '#,##0'

    ws['A5'] = 'Số hóa đơn'
    ws['B5'] = int(tong_hoa_don or 0)

    ws['A6'] = 'Tiền mặt'
    ws['B6'] = float(tien_mat or 0)
    ws['B6'].number_format = '#,##0'

    ws['A7'] = 'Chuyển khoản / QR'
    ws['B7'] = float(chuyen_khoan_qr or 0)
    ws['B7'].number_format = '#,##0'

    for row in range(4, 8):
        ws[f'A{row}'].font = bold

    headers = [
        'STT',
        'Mã hóa đơn',
        'Ngày',
        'Bàn',
        'Khách hàng',
        'Tổng tiền',
        'Thanh toán',
        'Thu ngân',
    ]

    start_row = 9
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=title)
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')

    for i, hd in enumerate(ds_hoa_don, start=1):
        row = start_row + i
        ten_ban = hd.ban.ten_ban if hd.ban else '—'
        if hd.ban_gop:
            ten_ban = f'{ten_ban} (gộp: {hd.ban_gop})'

        ten_kh = (
            hd.khach_hang.ten_khach_hang
            if hd.khach_hang
            else '—'
        )
        ten_nv = (
            hd.nguoi_dung.ho_ten
            if hd.nguoi_dung
            else '—'
        )
        ngay = (
            hd.created_at.strftime('%d/%m/%Y %H:%M')
            if hd.created_at
            else '—'
        )

        values = [
            i,
            hd.ma_hoa_don,
            ngay,
            ten_ban,
            ten_kh,
            float(hd.tong_tien or 0),
            _nhan_thanh_toan(hd.phuong_thuc_thanh_toan),
            ten_nv,
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin
            if col == 6:
                cell.number_format = '#,##0'

    widths = [6, 16, 18, 20, 22, 14, 18, 18]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
