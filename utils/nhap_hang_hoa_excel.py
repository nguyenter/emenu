import re
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook

from models.hang_hoa import HangHoa
from models.nhom_hang import NhomHang


HEADER_ALIASES = {
    'ma_hang': [
        'mã hàng', 'ma hang', 'ma_hang', 'mahàng', 'mahang', 'mã',
    ],
    'ten_hang': [
        'tên hàng', 'ten hang', 'ten_hang', 'tên', 'tenhang',
    ],
    'nhom_hang': [
        'nhóm hàng', 'nhom hang', 'nhom_hang', 'nhóm', 'nhomhang',
        'tên nhóm', 'ten nhom', 'mã nhóm', 'ma nhom',
    ],
    'don_vi_tinh': [
        'đơn vị tính', 'don vi tinh', 'don_vi_tinh', 'đơn vị', 'dvt',
    ],
    'gia_ban': [
        'giá bán', 'gia ban', 'gia_ban', 'giá', 'gia',
    ],
    'loai_thuc_don': [
        'loại thực đơn', 'loai thuc don', 'loai_thuc_don',
        'loại', 'loai',
    ],
}


def _chuan_hoa_tieu_de(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = text.replace('_', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text


def _map_cot(header_row):
    """Map index cột theo tiêu đề Excel."""
    mapping = {}
    for idx, cell in enumerate(header_row):
        title = _chuan_hoa_tieu_de(cell)
        if not title:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if title in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _doc_o(row, index):
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def _map_loai_thuc_don(value):
    if value is None:
        return None
    text = str(value).strip().upper()
    text = text.replace(' ', '_').replace('-', '_')

    aliases = {
        'MON_AN': 'MON_AN',
        'MONAN': 'MON_AN',
        'MÓN_ĂN': 'MON_AN',
        'MÓN ĂN': 'MON_AN',
        'MON AN': 'MON_AN',
        'DO_UONG': 'DO_UONG',
        'DOUONG': 'DO_UONG',
        'ĐỒ_UỐNG': 'DO_UONG',
        'ĐỒ UỐNG': 'DO_UONG',
        'DO UONG': 'DO_UONG',
        'COMBO': 'COMBO',
    }

    # So khớp không dấu / có dấu đơn giản
    raw = str(value).strip().lower()
    if raw in ('món ăn', 'mon an', 'mónăn'):
        return 'MON_AN'
    if raw in ('đồ uống', 'do uong', 'đồuống'):
        return 'DO_UONG'
    if raw == 'combo':
        return 'COMBO'

    return aliases.get(text)


def _parse_gia(value):
    if value is None or value == '':
        raise InvalidOperation('Thiếu giá bán')
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    text = str(value).strip()
    text = (
        text.replace('đ', '')
        .replace('Đ', '')
        .replace(' ', '')
        .replace('VND', '')
        .replace('vnd', '')
    )

    # 45.000 hoặc 45,000 → nghìn VNĐ
    if re.match(r'^\d{1,3}([.]\d{3})+$', text):
        text = text.replace('.', '')
    elif re.match(r'^\d{1,3}([,]\d{3})+$', text):
        text = text.replace(',', '')
    else:
        text = text.replace(',', '.')

    return Decimal(text)


def _tim_nhom_hang(ten_hoac_ma, chi_nhanh_id=None):
    if not ten_hoac_ma:
        return None
    key = str(ten_hoac_ma).strip()

    query = NhomHang.query
    if chi_nhanh_id is not None:
        query = query.filter(NhomHang.chi_nhanh_id == chi_nhanh_id)

    nhom = query.filter(
        (NhomHang.ten_nhom == key) | (NhomHang.ma_nhom == key)
    ).first()

    if nhom:
        return nhom

    # Không phân biệt hoa thường
    for item in query.all():
        if (
            (item.ten_nhom and item.ten_nhom.strip().lower() == key.lower())
            or (item.ma_nhom and item.ma_nhom.strip().lower() == key.lower())
        ):
            return item
    return None


def tao_file_mau_excel():
    """Tạo workbook mẫu để người dùng tải về."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'HangHoa'
    ws.append([
        'Mã hàng',
        'Tên hàng',
        'Nhóm hàng',
        'Đơn vị tính',
        'Giá bán',
        'Loại thực đơn',
    ])
    ws.append([
        'HH001',
        'Cơm gà',
        'Món chính',
        'Phần',
        45000,
        'MON_AN',
    ])
    ws.append([
        'DU001',
        'Trà đá',
        'Đồ uống',
        'Ly',
        10000,
        'DO_UONG',
    ])
    return wb


def nhap_hang_hoa_tu_excel(file_storage, chi_nhanh_id, db_session):
    """
    Đọc file Excel và thêm/cập nhật hàng hóa.
    Trả về dict: thanh_cong, bo_qua, loi (list message).
    """
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {
            'thanh_cong': 0,
            'cap_nhat': 0,
            'bo_qua': 0,
            'loi': ['File Excel trống'],
        }

    mapping = _map_cot(rows[0])
    bat_buoc = ['ma_hang', 'ten_hang', 'gia_ban', 'loai_thuc_don']
    thieu = [f for f in bat_buoc if f not in mapping]
    if thieu:
        return {
            'thanh_cong': 0,
            'cap_nhat': 0,
            'bo_qua': 0,
            'loi': [
                'Thiếu cột bắt buộc: '
                + ', '.join(thieu)
                + '. Cần có: Mã hàng, Tên hàng, Nhóm hàng, '
                'Đơn vị tính, Giá bán, Loại thực đơn'
            ],
        }

    ket_qua = {
        'thanh_cong': 0,
        'cap_nhat': 0,
        'bo_qua': 0,
        'loi': [],
    }

    for so_dong, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue

        ma_hang = _doc_o(row, mapping.get('ma_hang'))
        ten_hang = _doc_o(row, mapping.get('ten_hang'))
        nhom_raw = _doc_o(row, mapping.get('nhom_hang'))
        don_vi = _doc_o(row, mapping.get('don_vi_tinh'))
        gia_raw = _doc_o(row, mapping.get('gia_ban'))
        loai_raw = _doc_o(row, mapping.get('loai_thuc_don'))

        if not ma_hang or not ten_hang:
            ket_qua['loi'].append(
                f'Dòng {so_dong}: thiếu mã hàng hoặc tên hàng'
            )
            ket_qua['bo_qua'] += 1
            continue

        ma_hang = str(ma_hang).strip()
        ten_hang = str(ten_hang).strip()

        try:
            gia_ban = _parse_gia(gia_raw)
        except (InvalidOperation, ValueError):
            ket_qua['loi'].append(
                f'Dòng {so_dong}: giá bán không hợp lệ ({gia_raw})'
            )
            ket_qua['bo_qua'] += 1
            continue

        loai = _map_loai_thuc_don(loai_raw)
        if not loai:
            ket_qua['loi'].append(
                f'Dòng {so_dong}: loại thực đơn không hợp lệ '
                f'({loai_raw}). Dùng MON_AN / DO_UONG / COMBO'
            )
            ket_qua['bo_qua'] += 1
            continue

        nhom_hang_id = None
        if nhom_raw:
            nhom = _tim_nhom_hang(nhom_raw, chi_nhanh_id)
            if not nhom:
                ket_qua['loi'].append(
                    f'Dòng {so_dong}: không tìm thấy nhóm hàng "{nhom_raw}"'
                )
                ket_qua['bo_qua'] += 1
                continue
            nhom_hang_id = nhom.id

        hang = HangHoa.query.filter_by(ma_hang=ma_hang).first()
        if hang:
            hang.ten_hang = ten_hang
            hang.nhom_hang_id = nhom_hang_id
            hang.don_vi_tinh = str(don_vi).strip() if don_vi else hang.don_vi_tinh
            hang.gia_ban = gia_ban
            hang.loai_thuc_don = loai
            hang.trang_thai = True
            ket_qua['cap_nhat'] += 1
        else:
            hang = HangHoa(
                ma_hang=ma_hang,
                ten_hang=ten_hang,
                nhom_hang_id=nhom_hang_id,
                don_vi_tinh=str(don_vi).strip() if don_vi else None,
                gia_ban=gia_ban,
                loai_thuc_don=loai,
                trang_thai=True,
                chi_nhanh_id=chi_nhanh_id,
            )
            db_session.add(hang)
            ket_qua['thanh_cong'] += 1

    if ket_qua['thanh_cong'] or ket_qua['cap_nhat']:
        db_session.commit()

    return ket_qua
